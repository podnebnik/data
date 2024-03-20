#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 16 03:52:16 2023

@author: ziga
"""

import pandas as pd
import glob
import string
import openpyxl as xl

# # FIRST 100 excel columns
# # Get a list of all uppercase letters in the alphabet
# uppercase_letters = list(string.ascii_uppercase)

# # Generate a list of column names up to column CV (100th column)
# column_names = []
# for i in range(0, 100):
#     if i < 26:
#         column_names.append(uppercase_letters[i])
#     else:
#         first_letter = uppercase_letters[i // 26 - 1]
#         second_letter = uppercase_letters[i % 26]
#         column_names.append(first_letter + second_letter)

# print(column_names)

y = 1
for year in range(1986,2022):
    print(year)
    # Use a wildcard to match files with similar names
    file_pattern = './SVN_20230315/SVN_2023_{:04d}_*.xlsx'.format(year)

    # Use the glob module to find all files that match the pattern
    files = glob.glob(file_pattern)
    
    # Loop through each file and read in the data using pandas
    for file in files:
        print(file)
        wb1 = xl.load_workbook(file)
        ws1 = wb1.worksheets[56]
  
    # opening the destination excel file 
    filename1 ="emissions_historical_2023.xlsx"
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2.active

  
    # calculate total number of rows and 
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column
      
    # copying the cell values from source 
    # excel file to destination excel file
    for i in range (1, mr + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = 6+i, column = 10)
  
        # writing the read value to destination excel file
        # print(c.value)
        ws2.cell(row = i+1, column = y+1).value = c.value
      
    # saving the destination excel file
    wb2.save(filename1)
    
    y += 1

















#     # Read in the destination Excel file
#     destination_file = 'emissions_historical_2023_test.xlsx'
#     xl_destination = pd.ExcelFile(destination_file)
    
#     # Write the copied data to a specified sheet in the destination file
#     destination_sheet_name = 'Sheet1'
#     destination_writer = pd.ExcelWriter(destination_file, engine='openpyxl') 
#     destination_writer.book = load_workbook(destination_file)
    
#     # If the destination sheet exists, remove it before writing the new data
#     if destination_sheet_name in destination_writer.book.sheetnames:
#         idx = destination_writer.book.sheetnames.index(destination_sheet_name)
#         destination_writer.book.remove(destination_writer.book.worksheets[idx])
    
#     # Convert the copied data to a pandas DataFrame and write it to the destination file
#     copied_data_df = pd.DataFrame(copied_data, columns=[column_to_copy])
#     copied_data_df.to_excel(destination_writer, sheet_name=destination_sheet_name, startrow=1, startcol=i+1, index=False)

    
#     # Save the changes and close the writer object
#     destination_writer.save()
#     destination_writer.close()
       
